#!python
from model.model import Place, Direction, Item
from time import sleep
from openpyxl import load_workbook


def load_game():
    """Loads the game spreadsheet with details of places and moves"""
    wb = load_workbook(filename="game_files/nasty_house.xlsx")
    start_place = None
    places = {}
    sheet_names = wb.get_sheet_names()
    sheet = ""
    while sheet not in sheet_names:
        print("Please choose one of the following games:")

        for sheet_name in sheet_names:
            print(sheet_name)

        sheet=input("Please enter the name of the game you would like to play: ")
    sheet_ranges = wb.get_sheet_by_name(sheet)
    for row in sheet_ranges.iter_rows(row_offset = 0):
        command = row[0].value
        if command == "PLACE":
            place_id = row[1].value
            place_name = row[2].value
            place_description = row[3].value
            place_death = row[4].value == 1
            place_won = row[5].value == 1
            place = Place(place_name, place_description, place_death, place_won)
            if start_place is None:
                start_place = place
            places[place_id] = place
        elif command == "MOVE":
            direction = row[1].value
            start = row[2].value
            end = row[3].value
            places[start].add_place(Direction[direction] , places[end])
        elif command == "ITEM":
            item_name = row[1].value
            item_description = row[2].value
            place_found = places[row[3].value]
            place_used = places[row[4].value]
            destination = places[row[5].value]
            item = Item (item_name, item_description, place_found, place_used, destination)
            place_found.add_item(item)
            
    
    
    
    inventory = []
    return (start_place, inventory)

def display_room_info (current_place):
    print ("You are in the "+ current_place.name)
    print (current_place.description)

def start_game(start_place, inventory):
    current_place = start_place

    while True:
        display_room_info(current_place)
        command = input("?")
        command = command.lower()
        new_place = None
        if command == "n":
            new_place = current_place.move_direction(Direction.north)
        elif command == "e":
            new_place = current_place.move_direction(Direction.east)
        elif command == "s":
            new_place = current_place.move_direction(Direction.south)
        elif command == "w":
            new_place = current_place.move_direction(Direction.west)
        elif command == "look":
            if current_place.items:
                print ("You can see ")
                for item in current_place.items:
                    print (item.name)
            else:
                print ("There is nothing here")
        elif command.startswith("get "):
            thing_name = command[4:]
            thing = None
            for item in current_place.items:
                if item.name.lower() == thing_name.lower():
                    thing = item
                    break
            if thing is None:
                print ("You can't see one of those!")
            else:
                print("It is in your inventory\n")
                inventory.append(thing)
                current_place.items.remove(thing)
        elif command.startswith("use "):
            thing_name = command[4:]
            thing = None
            for item in inventory:
                if item.name.lower() == thing_name.lower():
                    thing = item
                    break
            if thing is None:
                print ("You haven't got one of those!")
            elif thing.can_use_here(current_place):
                new_place = thing.object_destination
                inventory.remove(thing)
            else:
                print("You cant use the "+ thing.name + " here")
        elif command.startswith("examine "):
            thing_name = command[8:]
            thing = None
            for item in inventory:
                if item.name.lower() == thing_name.lower():
                    thing = item
                    break
            if thing is None:
                print ("You don't have one of those!")
            else:
                print (thing.description + "\n")
        elif command == "inventory":
            if inventory:
                print ("you have:\n")
                for item in inventory:
                    print ("\t" + item.name)
                print("\n")
            else:
                print ("EMPTY INVENTORY\n")

        else:
            print("What?")
        if new_place is not None:
            current_place = new_place
        if current_place.won:
            display_room_info(current_place)
            sleep(2)
            print ("\n\nWell done! You won!")
            sleep(5)
            exit(0)
        if current_place.death:
            display_room_info(current_place)
            sleep(2)
            print ("\n\nYou died. The game is over")
            sleep(5)
            exit(0)


if __name__ == "__main__":
    """Load the game, run it"""

    
    (start_place, inventory) = load_game()
    start_game(start_place, inventory)
    
